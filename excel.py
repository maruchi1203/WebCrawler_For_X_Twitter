from openpyxl import load_workbook


def convert_to_num(num: str):
    if num == "" or num is None:
        return 0

    unit = num[-1]

    if unit == "만":
        unit = 10000
    elif unit == "천":
        unit = 1000
    elif unit == "K":
        unit = 1000
    elif unit == "M":
        unit = 1000000
    else:
        return int(num)

    ret_value = float(num[:-1]) * unit

    return int(ret_value)


class ExcelEdit:
    def __init__(self, file_url: str):
        self.__file_url = file_url

    def extract_file(self, data: dict):
        workbook = load_workbook(self.__file_url, data_only=True)
        worksheet = workbook.active

        keys = list(data.keys())
        values = list(data.values())

        for i in range(len(keys)):
            rowidx = 1
            while True:
                if worksheet.cell(row=rowidx, column=1).value is not None and worksheet.cell(row=rowidx, column=3).value != keys[i]:
                    rowidx += 1
                    continue
                else:
                    worksheet.cell(row=rowidx, column=1, value=values[i]["날짜"])
                    worksheet.cell(row=rowidx, column=2, value=values[i]["내용"])
                    worksheet.cell(row=rowidx, column=3, value=keys[i])
                    worksheet.cell(row=rowidx, column=4, value=convert_to_num(values[i]["리포스트"]))
                    worksheet.cell(row=rowidx, column=5, value=convert_to_num(values[i]["좋아요"]))
                    worksheet.cell(row=rowidx, column=6, value=convert_to_num(values[i]["조회수"]))
                    break

        workbook.save(self.__file_url)
